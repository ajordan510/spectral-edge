"""
Excel Output Module for Batch Processing

This module handles exporting batch PSD processing results to Excel format.
Each event is written to a separate sheet with PSD data and summary statistics.

Author: SpectralEdge Development Team
Date: 2026-02-02
"""

import numpy as np
import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from spectral_edge.batch.output_utils import sanitize_filename_component as _sanitize_filename_component

logger = logging.getLogger(__name__)


def _build_event_dataframe(event_results: Dict) -> Optional[pd.DataFrame]:
    """Build an event dataframe by aligning channel PSDs on frequency (outer join)."""
    merged_df = None
    for channel_id, result_data in event_results.items():
        frequencies = np.asarray(result_data['frequencies'])
        psd_values = np.asarray(result_data['psd'])
        if frequencies.size == 0 or psd_values.size == 0:
            continue
        length = min(frequencies.size, psd_values.size)
        channel_df = pd.DataFrame(
            {
                "Frequency (Hz)": frequencies[:length],
                channel_id: psd_values[:length],
            }
        ).drop_duplicates(subset=["Frequency (Hz)"], keep="first")
        if merged_df is None:
            merged_df = channel_df
        else:
            merged_df = merged_df.merge(channel_df, on="Frequency (Hz)", how="outer")

    if merged_df is None or merged_df.empty:
        return None
    return merged_df.sort_values("Frequency (Hz)").reset_index(drop=True)


def _describe_filter(config: 'BatchConfig') -> str:
    """Create deterministic filter summary text for exports."""
    if config.filter_config.enabled:
        hp = getattr(config.filter_config, "user_highpass_hz", None)
        lp = getattr(config.filter_config, "user_lowpass_hz", None)
        if hp is None:
            hp = getattr(config.filter_config, "cutoff_low", None)
        if lp is None:
            lp = getattr(config.filter_config, "cutoff_high", None)
        hp_text = "baseline" if hp is None else f"{float(hp):g}"
        lp_text = "baseline" if lp is None else f"{float(lp):g}"
        return f"Baseline + user overrides (HP={hp_text} Hz, LP={lp_text} Hz)"
    return "Baseline only (HP 1.0 Hz, LP 0.45xFs)"


def export_to_excel(
    result: 'BatchProcessingResult',
    output_directory: str,
    config: 'BatchConfig',
    filename: str = "batch_psd_results.xlsx"
) -> str:
    """
    Export batch processing results to Excel file.
    
    Creates an Excel workbook where each event has its own sheet containing
    PSD data for all processed channels.
    
    Parameters:
    -----------
    result : BatchProcessingResult
        Batch processing result object containing channel_results
    output_directory : str
        Directory to save the Excel file
    config : BatchConfig
        Batch configuration object (for spacing conversion)
    filename : str, optional
        Name of the Excel file (default: "batch_psd_results.xlsx")
        
    Returns:
    --------
    str
        Full path to the saved Excel file
    
    Raises:
    -------
    IOError
        If file cannot be written
    """
    try:
        prefix = _sanitize_filename_component(getattr(config.output_config, "filename_prefix", ""))
        resolved_filename = filename or "batch_psd_results.xlsx"
        if prefix:
            resolved_filename = f"{prefix}_{resolved_filename}"
        output_path = str(Path(output_directory) / resolved_filename)
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Reorganize results by event
        from spectral_edge.batch.output_psd import apply_frequency_spacing

        events_data = {}
        for (flight_key, channel_key), event_dict in result.channel_results.items():
            for event_name, event_result in event_dict.items():
                if event_name not in events_data:
                    events_data[event_name] = {}
                channel_id = f"{flight_key}_{channel_key}"
                frequencies_out, psd_out = apply_frequency_spacing(
                    event_result['frequencies'],
                    event_result['psd'],
                    config.psd_config
                )
                events_data[event_name][channel_id] = {
                    'frequencies': frequencies_out,
                    'psd': psd_out,
                    'metadata': event_result['metadata']
                }
        
        # Create summary sheet
        _create_summary_sheet(wb, events_data, result, config)
        
        # Create a sheet for each event
        for event_name, event_results in events_data.items():
            _create_event_sheet(wb, event_name, event_results)
        
        # Save workbook
        wb.save(output_path)
        logger.info(f"Excel file saved: {output_path}")
        return output_path
        
    except Exception as e:
        logger.error(f"Failed to export to Excel: {str(e)}")
        raise


def _create_summary_sheet(wb: Workbook, events_data: Dict, result: 'BatchProcessingResult', config: 'BatchConfig') -> None:
    """
    Create summary sheet with overview of all events and channels.
    
    Parameters:
    -----------
    wb : Workbook
        Excel workbook object
    events_data : Dict
        Reorganized events data
    result : BatchProcessingResult
        Batch processing result object
    """
    ws = wb.create_sheet("Summary", 0)
    
    # Title
    ws['A1'] = "Batch PSD Processing Summary"
    ws['A1'].font = Font(size=14, bold=True)
    
    # Processing info
    row = 3
    ws[f'A{row}'] = "Processing Summary:"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    run_time = result.start_time.strftime("%Y-%m-%d %H:%M:%S") if result.start_time else ""
    ws[f'A{row}'] = "Run Timestamp:"
    ws[f'B{row}'] = run_time
    row += 1

    source_files = ", ".join([Path(p).name for p in config.source_files]) if config.source_files else ""
    ws[f'A{row}'] = "Source Files:"
    ws[f'B{row}'] = source_files
    row += 1

    ws[f'A{row}'] = "Total Channels:"
    ws[f'B{row}'] = len(result.channel_results)
    row += 1
    
    ws[f'A{row}'] = "Total Events:"
    ws[f'B{row}'] = len(events_data)
    row += 1
    
    ws[f'A{row}'] = "Errors:"
    ws[f'B{row}'] = len(result.errors)
    row += 1
    
    ws[f'A{row}'] = "Warnings:"
    ws[f'B{row}'] = len(result.warnings)
    row += 2

    ws[f'A{row}'] = "Processing Parameters:"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = "PSD Method:"
    ws[f'B{row}'] = config.psd_config.method
    row += 1
    ws[f'A{row}'] = "Window:"
    ws[f'B{row}'] = config.psd_config.window
    row += 1
    ws[f'A{row}'] = "df (Hz):"
    ws[f'B{row}'] = config.psd_config.desired_df
    row += 1
    ws[f'A{row}'] = "Overlap (%):"
    ws[f'B{row}'] = config.psd_config.overlap_percent
    row += 1
    ws[f'A{row}'] = "Efficient FFT:"
    ws[f'B{row}'] = "On" if config.psd_config.use_efficient_fft else "Off"
    row += 1
    ws[f'A{row}'] = "Frequency Spacing:"
    ws[f'B{row}'] = config.psd_config.frequency_spacing
    row += 1
    filt_desc = _describe_filter(config)
    ws[f'A{row}'] = "Filter:"
    ws[f'B{row}'] = filt_desc
    row += 2
    
    # Events summary
    ws[f'A{row}'] = "Events Processed:"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    # Create table header
    ws[f'A{row}'] = "Event Name"
    ws[f'B{row}'] = "Channels Processed"
    ws[f'C{row}'] = "Sources"
    
    for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}']]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    row += 1
    
    # Fill in event data
    for event_name, event_results in events_data.items():
        ws[f'A{row}'] = event_name
        
        # Count channels
        total_channels = len(event_results)
        ws[f'B{row}'] = total_channels
        
        # List channel IDs
        channels = ", ".join(list(event_results.keys())[:5])  # Show first 5
        if len(event_results) > 5:
            channels += "..."
        ws[f'C{row}'] = channels
        
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40

    # RMS summary table
    row += 1
    ws[f'A{row}'] = "RMS Summary:"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    headers = ["Flight", "Channel", "Event", "RMS", "3-Sigma RMS"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    row += 1

    for (flight_key, channel_key), event_dict in result.channel_results.items():
        for event_name, event_result in event_dict.items():
            rms_value = event_result.get('metadata', {}).get('rms')
            ws.cell(row=row, column=1, value=flight_key)
            ws.cell(row=row, column=2, value=channel_key)
            ws.cell(row=row, column=3, value=event_name)
            ws.cell(row=row, column=4, value=None if rms_value is None else float(rms_value))
            ws.cell(row=row, column=5, value=None if rms_value is None else float(3.0 * rms_value))
            row += 1


def _create_event_sheet(wb: Workbook, event_name: str, event_results: Dict) -> None:
    """
    Create a sheet for a specific event with PSD data.
    
    Parameters:
    -----------
    wb : Workbook
        Excel workbook object
    event_name : str
        Name of the event
    event_results : Dict
        Results for this event (channel_id -> {frequencies, psd, metadata})
    """
    # Clean sheet name (Excel has restrictions)
    sheet_name = event_name[:31]  # Max 31 characters
    ws = wb.create_sheet(sheet_name)
    
    # Title
    ws['A1'] = f"Event: {event_name}"
    ws['A1'].font = Font(size=12, bold=True)
    
    df = _build_event_dataframe(event_results)
    if df is None:
        logger.warning(f"No data for event: {event_name}")
        return
    
    # Write to sheet starting at row 3
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=3):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Format header row
            if r_idx == 3:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


